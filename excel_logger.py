import os
import traceback
from datetime import datetime

import openpyxl
from openpyxl import load_workbook, Workbook


class WorkflowStopRequested(Exception):
    """指示工作流应在此节点后停止的特殊异常。"""
    pass


def _col_letter_to_index(letter):
    """将 Excel 列字母（如 'A', 'B', 'AA'）转换为 1 基索引（openpyxl 使用 1 基）。"""
    letter = (letter or "").strip().upper()
    if not letter:
        return 0  # 0 表示"跳过该列"
    index = 0
    for char in letter:
        if not 'A' <= char <= 'Z':
            raise ValueError(f"列字母 '{letter}' 中包含无效字符 '{char}'。")
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index


def _write_cell(ws, row, col_letter, value):
    """按列字母写入单元格；列字母为空则跳过。"""
    col = _col_letter_to_index(col_letter)
    if col <= 0:
        return
    ws.cell(row=row, column=col, value=value)


class ExcelAutoLogger:
    """
    QC.LogToExcel

    将 Prompt 及其生成参数写入 Excel 指定列。
    - 使用 Excel 列字母（A/B/.../AA）指定每个字段写入的列，留空则跳过该字段。
    - mode='auto_increment' 时 start_row_number 每次运行后自动 +1；
      超出 end_row_number 时抛出 WorkflowStopRequested 停止工作流。
    """

    def __init__(self):
        pass

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "prompt": ("STRING", {"multiline": True, "forceInput": True}),
                "excel_path": ("STRING", {"default": "prompt_log.xlsx"}),
                "sheet_name": ("STRING", {"default": "Prompts"}),
                "mode": (["auto_increment", "manual"], {"default": "auto_increment"}),
                "start_row_number": ("INT", {"default": 2, "min": 1, "max": 1000000, "step": 1, "control_after_generate": True}),
                "end_row_number": ("INT", {"default": -1, "min": -1, "max": 1000000, "step": 1, "description": "-1 表示无上限"}),
                "prompt_column": ("STRING", {"default": "C", "description": "写入 prompt 的列字母，留空则跳过"}),
                "id_column": ("STRING", {"default": "A", "description": "写入自增ID的列字母，留空则跳过"}),
                "timestamp_column": ("STRING", {"default": "B", "description": "写入时间戳的列字母，留空则跳过"}),
                "write_header": (["yes", "no"], {"default": "yes", "description": "新建工作表时是否写入表头行"}),
            },
            "optional": {
                "negative_prompt": ("STRING", {"multiline": True, "forceInput": True}),
                "negative_column": ("STRING", {"default": "D"}),
                "seed": ("INT", {"forceInput": True}),
                "seed_column": ("STRING", {"default": "E"}),
                "model_name": ("STRING", {"forceInput": True}),
                "model_column": ("STRING", {"default": "F"}),
                "steps": ("INT", {"forceInput": True}),
                "steps_column": ("STRING", {"default": "G"}),
                "cfg": ("FLOAT", {"forceInput": True}),
                "cfg_column": ("STRING", {"default": "H"}),
                "image_path": ("STRING", {"forceInput": True}),
                "image_path_column": ("STRING", {"default": "I"}),
                "image_name": ("STRING", {"forceInput": True}),
                "image_name_column": ("STRING", {"default": "J"}),
                "url": ("STRING", {"forceInput": True}),
                "url_column": ("STRING", {"default": "K"}),
            }
        }

    RETURN_TYPES = ("STRING", "INT")
    RETURN_NAMES = ("prompt", "row_number")
    FUNCTION = "log_to_excel"
    CATEGORY = "ExcelUtils"
    OUTPUT_NODE = True

    def log_to_excel(self, prompt, excel_path, sheet_name, mode,
                     start_row_number, end_row_number,
                     prompt_column, id_column, timestamp_column, write_header,
                     negative_prompt=None, negative_column="D",
                     seed=None, seed_column="E",
                     model_name=None, model_column="F",
                     steps=None, steps_column="G",
                     cfg=None, cfg_column="H",
                     image_path=None, image_path_column="I",
                     image_name=None, image_name_column="J",
                     url=None, url_column="K"):
        try:
            if not excel_path:
                raise ValueError("excel_path 不能为空。")

            # --- 打开/创建工作簿 ---
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
            else:
                parent = os.path.dirname(excel_path)
                if parent and not os.path.exists(parent):
                    os.makedirs(parent, exist_ok=True)
                wb = Workbook()
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])

            # --- 获取/创建工作表 ---
            new_sheet = sheet_name not in wb.sheetnames
            if new_sheet:
                ws = wb.create_sheet(sheet_name)
                if write_header == "yes":
                    header_map = [
                        (id_column, "ID"),
                        (timestamp_column, "Timestamp"),
                        (prompt_column, "Prompt"),
                        (negative_column, "Negative Prompt"),
                        (seed_column, "Seed"),
                        (model_column, "Model"),
                        (steps_column, "Steps"),
                        (cfg_column, "CFG"),
                        (image_path_column, "Image Path"),
                        (image_name_column, "Image Name"),
                        (url_column, "URL"),
                    ]
                    for col_letter, title in header_map:
                        _write_cell(ws, 1, col_letter, title)
            else:
                ws = wb[sheet_name]

            # --- 确定写入行号 ---
            current_row = int(start_row_number)
            # prompt_column 决定"空行"检测基准；若留空则退回到 A 列
            probe_col = _col_letter_to_index(prompt_column) or 1

            if mode == "auto_increment":
                row = current_row
                while ws.cell(row=row, column=probe_col).value is not None:
                    row += 1
                current_row = row

            # --- 结束行检查 ---
            if end_row_number != -1 and current_row > int(end_row_number):
                msg = f"已到达结束行 ({end_row_number})，当前行 {current_row}。工作流停止。"
                print(f"[QC.LogToExcel] {msg}")
                raise WorkflowStopRequested(msg)

            # --- 写入数据 ---
            _write_cell(ws, current_row, id_column, current_row - 1)
            _write_cell(ws, current_row, timestamp_column,
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            _write_cell(ws, current_row, prompt_column, prompt)
            _write_cell(ws, current_row, negative_column, negative_prompt or "")
            _write_cell(ws, current_row, seed_column, seed)
            _write_cell(ws, current_row, model_column, model_name or "")
            _write_cell(ws, current_row, steps_column, steps)
            _write_cell(ws, current_row, cfg_column, cfg)
            _write_cell(ws, current_row, image_path_column, image_path or "")
            _write_cell(ws, current_row, image_name_column, image_name or "")
            _write_cell(ws, current_row, url_column, url or "")

            wb.save(excel_path)
            print(f"[QC.LogToExcel] ✓ 写入 {excel_path} -> {sheet_name}!行{current_row}")

            # --- 计算下次起始行 ---
            next_start = current_row + 1
            if end_row_number != -1 and next_start > int(end_row_number):
                next_start = int(end_row_number) + 1

            return {
                "ui": {"start_row_number": [next_start]},
                "result": (prompt, current_row),
            }

        except WorkflowStopRequested:
            raise
        except Exception as e:
            print(f"[QC.LogToExcel] ✗ 写入错误: {e}")
            traceback.print_exc()
            return {"result": (prompt, -1)}


NODE_CLASS_MAPPINGS = {
    "QC.LogToExcel": ExcelAutoLogger,
    # 兼容旧工作流中的节点名
    "ExcelAutoLogger": ExcelAutoLogger,
}

NODE_DISPLAY_NAME_MAPPINGS = {
    "QC.LogToExcel": "QC.LogToExcel",
    "ExcelAutoLogger": "QC.LogToExcel (legacy)",
}

print("--- 加载自定义节点: QC.LogToExcel (ExcelUtils) v2.0 ---")
